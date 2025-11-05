"""
Excel Export Module
Export ECG analysis results to Excel with formatting and charts
"""

import os
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def export_to_excel(records, output_path=None):
    """
    Export ECG analysis records to Excel file
    
    Args:
        records: list - List of records from database
        output_path: str - Output file path (default: ecg_analysis_TIMESTAMP.xlsx)
    
    Returns:
        str - Path to exported file
    """
    
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")
    
    if output_path is None:
        output_path = f"ecg_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 1. Create Data Sheet
    ws_data = wb.create_sheet("ECG Data", 0)
    _create_data_sheet(ws_data, records)
    
    # 2. Create Summary Sheet
    ws_summary = wb.create_sheet("Summary", 1)
    _create_summary_sheet(ws_summary, records)
    
    # 3. Create Statistics Sheet
    ws_stats = wb.create_sheet("Statistics", 2)
    _create_statistics_sheet(ws_stats, records)
    
    # Save file
    wb.save(output_path)
    print(f"✓ Excel file exported: {output_path}")
    
    return output_path


def _create_data_sheet(ws, records):
    """Create data sheet with all records"""
    
    # Headers
    headers = [
        'File Name', 'Date & Time', 'Feature Set',
        'Heart Rate (bpm)', 'SDNN (ms)', 'RMSSD (ms)', 'pNN50 (%)', 'pNN20 (%)',
        'SD1 (ms)', 'SD2 (ms)', 'LF Power', 'HF Power', 'LF/HF Ratio'
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data rows
    for row_idx, record in enumerate(records, 2):
        metrics = record.get('metrics', record)  # Handle both formats
        
        values = [
            record.get('filename', 'N/A'),
            record.get('timestamp', record.get('upload_timestamp', 'N/A')),
            record.get('features', record.get('feature_set', 'standard')),
            metrics.get('heart_rate_bpm'),
            metrics.get('sdnn_ms'),
            metrics.get('rmssd_ms'),
            metrics.get('pnn50_percent'),
            metrics.get('pnn20_percent'),
            metrics.get('sd1_ms'),
            metrics.get('sd2_ms'),
            metrics.get('lf_power'),
            metrics.get('hf_power'),
            metrics.get('lf_hf_ratio'),
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            
            # Format numeric columns
            if col >= 4 and isinstance(value, (int, float)):
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
    
    # Adjust column widths
    column_widths = [20, 20, 15, 15, 12, 12, 12, 12, 12, 12, 12, 12, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"


def _create_summary_sheet(ws, records):
    """Create summary sheet with key statistics"""
    
    ws['A1'] = 'ECG Analysis Summary Report'
    ws['A1'].font = Font(bold=True, size=14, color="667eea")
    
    row = 3
    
    # Basic statistics
    ws[f'A{row}'] = 'Analysis Summary'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    
    stats = [
        ('Total Records', len(records)),
        ('Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
    ]
    
    for label, value in stats:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    row += 1
    
    # Metric statistics
    if records:
        ws[f'A{row}'] = 'Metric Statistics'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Calculate averages
        metrics_to_calc = [
            ('heart_rate_bpm', 'Heart Rate (bpm)'),
            ('sdnn_ms', 'SDNN (ms)'),
            ('rmssd_ms', 'RMSSD (ms)'),
            ('pnn50_percent', 'pNN50 (%)'),
            ('pnn20_percent', 'pNN20 (%)'),
            ('sd1_ms', 'SD1 (ms)'),
            ('sd2_ms', 'SD2 (ms)'),
            ('lf_power', 'LF Power'),
            ('hf_power', 'HF Power'),
            ('lf_hf_ratio', 'LF/HF Ratio'),
        ]
        
        # Headers
        ws[f'A{row}'] = 'Metric'
        ws[f'B{row}'] = 'Min'
        ws[f'C{row}'] = 'Max'
        ws[f'D{row}'] = 'Average'
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color="e8eaff", end_color="e8eaff", fill_type="solid")
        row += 1
        
        for metric_key, metric_label in metrics_to_calc:
            values = [
                r.get('metrics', r).get(metric_key)
                for r in records
                if r.get('metrics', r).get(metric_key) is not None
            ]
            
            if values:
                ws[f'A{row}'] = metric_label
                ws[f'B{row}'] = min(values)
                ws[f'C{row}'] = max(values)
                ws[f'D{row}'] = sum(values) / len(values)
                
                # Format numbers
                for col in ['B', 'C', 'D']:
                    ws[f'{col}{row}'].number_format = '0.00'
                
                row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15


def _create_statistics_sheet(ws, records):
    """Create detailed statistics sheet"""
    
    ws['A1'] = 'Detailed Statistics'
    ws['A1'].font = Font(bold=True, size=14, color="667eea")
    
    if not records:
        ws['A3'] = 'No data available'
        return
    
    row = 3
    
    # Distribution by feature set
    feature_sets = {}
    for record in records:
        fs = record.get('features', record.get('feature_set', 'unknown'))
        feature_sets[fs] = feature_sets.get(fs, 0) + 1
    
    ws[f'A{row}'] = 'Records by Feature Set'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    ws[f'A{row}'] = 'Feature Set'
    ws[f'B{row}'] = 'Count'
    for col in ['A', 'B']:
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = PatternFill(start_color="e8eaff", end_color="e8eaff", fill_type="solid")
    row += 1
    
    for fs, count in sorted(feature_sets.items()):
        ws[f'A{row}'] = fs
        ws[f'B{row}'] = count
        row += 1
    
    row += 2
    
    # Data quality metrics
    ws[f'A{row}'] = 'Data Quality'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    total = len(records)
    complete = sum(1 for r in records if all(
        r.get('metrics', r).get(k) is not None 
        for k in ['heart_rate_bpm', 'sdnn_ms', 'rmssd_ms', 'lf_hf_ratio']
    ))
    
    quality_metrics = [
        ('Total Records', total),
        ('Complete Records', complete),
        ('Completeness %', f"{(complete/total*100):.1f}%" if total > 0 else "0%"),
    ]
    
    for label, value in quality_metrics:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20


if __name__ == '__main__':
    # Test export
    test_records = [
        {
            'filename': 'test1.pdf',
            'timestamp': datetime.now().isoformat(),
            'features': 'standard',
            'metrics': {
                'heart_rate_bpm': 72.5,
                'sdnn_ms': 48.2,
                'rmssd_ms': 35.8,
                'pnn50_percent': 28.3,
                'pnn20_percent': 65.4,
                'sd1_ms': 25.3,
                'sd2_ms': 52.1,
                'lf_power': 425.3,
                'hf_power': 198.7,
                'lf_hf_ratio': 2.1
            }
        }
    ]
    
    if OPENPYXL_AVAILABLE:
        export_to_excel(test_records, 'test_export.xlsx')
        print("✓ Test export created: test_export.xlsx")
    else:
        print("⚠ openpyxl not installed")
